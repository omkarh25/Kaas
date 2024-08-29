from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Boolean, ForeignKey, desc, asc
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from pydantic import BaseModel
from datetime import datetime, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import logging
from typing import Optional, List
from openpyxl import load_workbook
import os
from threading import local

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Database setup
SQLALCHEMY_DATABASE_URL = "sqlite:///./kaas.db?check_same_thread=False"
engine = create_engine(SQLALCHEMY_DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Security
SECRET_KEY = "your-secret-key"  # Change this!
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# Create a thread-local object to store the database session
thread_local = local()

def get_db():
    if not hasattr(thread_local, "db"):
        thread_local.db = SessionLocal()
    try:
        yield thread_local.db
    finally:
        thread_local.db.close()

# Models
class User(Base):
    __tablename__ = "users"

    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    hashed_password = Column(String)
    is_admin = Column(Boolean, default=False)

class Asset(Base):
    __tablename__ = "assets"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, index=True)
    amount = Column(Float)
    department = Column(String)
    comments = Column(String)

class Transaction(Base):
    __tablename__ = "transactions"

    id = Column(Integer, primary_key=True, index=True)
    date = Column(DateTime, default=datetime.utcnow)
    description = Column(String)
    amount = Column(Float)
    payment_mode = Column(String)
    acc_id = Column(String)
    department = Column(String)
    comments = Column(String)
    category = Column(String)
    payment_mode_detail = Column(String)
    zoho_match = Column(Boolean, default=False)

# Pydantic models
class UserCreate(BaseModel):
    username: str
    password: str
    is_admin: bool = False

class UserInDB(BaseModel):
    id: int
    username: str
    is_admin: bool

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: str | None = None

class AssetCreate(BaseModel):
    name: str
    amount: float
    department: str
    comments: str

class AssetResponse(AssetCreate):
    id: int

class TransactionCreate(BaseModel):
    date: datetime
    description: str
    amount: float
    payment_mode: str
    acc_id: str
    department: str
    comments: Optional[str] = None
    category: str
    payment_mode_detail: str
    zoho_match: bool = False

class TransactionResponse(TransactionCreate):
    id: int

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# Load Excel data
def load_excel_data(db: Session):
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'Kaas.xlsx')
    workbook = load_workbook(filename=excel_path)

    # Load Assets
    sheet = workbook['Assets']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        asset = Asset(
            name=row[1],
            amount=row[2],
            department=row[3],
            comments=row[4]
        )
        db.add(asset)

    # Load Transactions
    sheet = workbook['Transactions(Past)']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction = Transaction(
            date=row[1],
            description=row[2],
            amount=row[3],
            payment_mode=row[4],
            acc_id=row[5],
            department=row[6],
            comments=row[7],
            category=row[8],
            payment_mode_detail=row[9],
            zoho_match=row[10] == 'Yes'
        )
        db.add(transaction)

    db.commit()
    logger.info("Excel data loaded successfully")

# Security functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def authenticate_user(db: Session, username: str, password: str):
    user = db.query(User).filter(User.username == username).first()
    if not user or not verify_password(password, user.hashed_password):
        return False
    return user

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.username == token_data.username).first()
    if user is None:
        raise credentials_exception
    return user

async def get_current_active_user(current_user: UserInDB = Depends(get_current_user)):
    return current_user

# Routes
@app.on_event("startup")
async def startup_event():
    db = SessionLocal()
    load_excel_data(db)
    db.close()

@app.post("/users/")
async def create_user(user: UserCreate):
    # Your user creation logic here
    # ...

    # Return a dictionary with the created user's information
    return {"username": user.username, "is_admin": user.is_admin}

@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    logger.info(f"User logged in: {user.username}")
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/assets/", response_model=AssetResponse)
def create_asset(asset: AssetCreate, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to create assets")
    db_asset = Asset(**asset.dict())
    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)
    logger.info(f"Created asset: {db_asset.name}")
    return db_asset

@app.get("/assets/", response_model=List[AssetResponse])
def read_assets(
    skip: int = 0,
    limit: int = 100,
    sort_by: str = "id",
    sort_order: str = "asc",
    name: Optional[str] = None,
    department: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    query = db.query(Asset)
    
    # Apply filters
    if name:
        query = query.filter(Asset.name.ilike(f"%{name}%"))
    if department:
        query = query.filter(Asset.department == department)
    
    # Apply sorting
    if sort_order == "asc":
        query = query.order_by(asc(getattr(Asset, sort_by)))
    else:
        query = query.order_by(desc(getattr(Asset, sort_by)))
    
    assets = query.offset(skip).limit(limit).all()
    logger.info(f"Retrieved {len(assets)} assets")
    return assets

@app.get("/assets/{asset_id}", response_model=AssetResponse)
def read_asset(asset_id: int, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    db_asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if db_asset is None:
        logger.warning(f"Asset not found: {asset_id}")
        raise HTTPException(status_code=404, detail="Asset not found")
    logger.info(f"Retrieved asset: {db_asset.name}")
    return db_asset

@app.put("/assets/{asset_id}", response_model=AssetResponse)
def update_asset(asset_id: int, asset: AssetCreate, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to update assets")
    db_asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if db_asset is None:
        logger.warning(f"Asset not found: {asset_id}")
        raise HTTPException(status_code=404, detail="Asset not found")
    for key, value in asset.dict().items():
        setattr(db_asset, key, value)
    db.commit()
    db.refresh(db_asset)
    logger.info(f"Updated asset: {db_asset.name}")
    return db_asset

@app.delete("/assets/{asset_id}")
def delete_asset(asset_id: int, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to delete assets")
    db_asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if db_asset is None:
        logger.warning(f"Asset not found: {asset_id}")
        raise HTTPException(status_code=404, detail="Asset not found")
    db.delete(db_asset)
    db.commit()
    logger.info(f"Deleted asset: {asset_id}")
    return {"message": "Asset deleted successfully"}

@app.post("/transactions/", response_model=TransactionResponse)
def create_transaction(transaction: TransactionCreate, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to create transactions")
    db_transaction = Transaction(**transaction.dict())
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    logger.info(f"Created transaction: {db_transaction.description}")
    return db_transaction

@app.get("/transactions/", response_model=List[TransactionResponse])
def read_transactions(
    skip: int = 0,
    limit: int = 100,
    sort_by: str = "date",
    sort_order: str = "desc",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    category: Optional[str] = None,
    department: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    query = db.query(Transaction)
    
    # Apply filters
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if category:
        query = query.filter(Transaction.category == category)
    if department:
        query = query.filter(Transaction.department == department)
    
    # Apply sorting
    if sort_order == "asc":
        query = query.order_by(asc(getattr(Transaction, sort_by)))
    else:
        query = query.order_by(desc(getattr(Transaction, sort_by)))
    
    transactions = query.offset(skip).limit(limit).all()
    logger.info(f"Retrieved {len(transactions)} transactions")
    return transactions

@app.get("/transactions/{transaction_id}", response_model=TransactionResponse)
def read_transaction(transaction_id: int, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    db_transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if db_transaction is None:
        logger.warning(f"Transaction not found: {transaction_id}")
        raise HTTPException(status_code=404, detail="Transaction not found")
    logger.info(f"Retrieved transaction: {db_transaction.description}")
    return db_transaction

@app.put("/transactions/{transaction_id}", response_model=TransactionResponse)
def update_transaction(transaction_id: int, transaction: TransactionCreate, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to update transactions")
    db_transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if db_transaction is None:
        logger.warning(f"Transaction not found: {transaction_id}")
        raise HTTPException(status_code=404, detail="Transaction not found")
    for key, value in transaction.dict().items():
        setattr(db_transaction, key, value)
    db.commit()
    db.refresh(db_transaction)
    logger.info(f"Updated transaction: {db_transaction.description}")
    return db_transaction

@app.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, current_user: UserInDB = Depends(get_current_active_user), db: Session = Depends(get_db)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not authorized to delete transactions")
    db_transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if db_transaction is None:
        logger.warning(f"Transaction not found: {transaction_id}")
        raise HTTPException(status_code=404, detail="Transaction not found")
    db.delete(db_transaction)
    db.commit()
    logger.info(f"Deleted transaction: {transaction_id}")
    return {"message": "Transaction deleted successfully"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)