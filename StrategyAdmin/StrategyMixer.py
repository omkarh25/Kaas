import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Margin and lot sizes
margin_lot_data = {
    'nf': {'margin_per_lot': 70000, 'lot_size': 50},
    'bnf': {'margin_per_lot': 22000, 'lot_size': 15},
    'fnf': {'margin_per_lot': 24000, 'lot_size': 40},
    'midcp': {'margin_per_lot': 24000, 'lot_size': 75},
    'sensex': {'margin_per_lot': 18000, 'lot_size': 10},
    'amipy': {'margin_per_lot': 70000, 'lot_size': 50},
    'overnight': {'margin_per_lot': 70000, 'lot_size': 50}
}

# Initial Risk Profiles
risk_profile = {
    "AmiPy": 250000,  # rs/lot
    "MPWizard": 0.01,  # % of total capital
    "OvernightFutures": 200000,  # rs/lot
    "ExpiryTrader": 0.01,  # % of total capital
    "Stocks": 0.01  # % of total capital
}

CAPITAL = 2500000.00  # Starting capital in rs

# Function to read the Excel file and return a dictionary of DataFrames
def read_strategy_logs(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    strategy_data = {}
    for strategy in ['AmiPy', 'MPWizard', 'ExpiryTrader', 'OvernightFutures', 'Stocks']:
        if strategy in wb.sheetnames:
            df = pd.read_excel(file_path, sheet_name=strategy)
            # Ensure entry_time and exit_time are in datetime format
            df['entry_time'] = pd.to_datetime(df['entry_time'])
            df['exit_time'] = pd.to_datetime(df['exit_time'])
            strategy_data[strategy] = df
    return strategy_data


# Function to calculate margins based on the strategy
def calculate_margins(strategy_data):
    for strategy, df in strategy_data.items():
        if strategy == "AmiPy":
            df['margin'] = df['qty'] * margin_lot_data['amipy']['margin_per_lot']
        elif strategy == "MPWizard":
            df['margin'] = df['qty'] * df['entry_price']
        elif strategy == "ExpiryTrader":
            # Implement a check to extract relevant part of the trading_symbol
            def extract_margin(row):
                # Identify the symbol (e.g., nf, bnf, etc.) from the trading symbol
                trading_symbol = row['trading_symbol'].lower()
                if 'nifty' in trading_symbol:
                    symbol_key = 'nf'
                elif 'banknifty' in trading_symbol:
                    symbol_key = 'bnf'
                elif 'finnifty' in trading_symbol:
                    symbol_key = 'fnf'
                elif 'midcp' in trading_symbol:  # Update this as per your symbols
                    symbol_key = 'midcp'
                elif 'sensex' in trading_symbol:
                    symbol_key = 'sensex'
                else:
                    symbol_key = 'default'  # Handle or log unexpected symbols

                # Calculate margin using the identified symbol
                return row['qty'] * margin_lot_data[symbol_key]['margin_per_lot']

            df['margin'] = df.apply(extract_margin, axis=1)
        elif strategy == "OvernightFutures":
            df['margin'] = df['qty'] * margin_lot_data['overnight']['margin_per_lot']
        elif strategy == "Stocks":
            df['margin'] = df['qty'] * df['entry_price']
    return strategy_data


            
def calculate_equity(strategy_data):
    min_dates = []
    max_dates = []
    for df in strategy_data.values():
        min_dates.append(df['entry_time'].min().date())  # Ensure min date is datetime.date
        max_dates.append(df['exit_time'].max().date())  # Ensure max date is datetime.date

    min_date = min(min_dates)
    max_date = max(max_dates)

    dates = pd.date_range(start=min_date, end=max_date, freq='D')
    equity_df = pd.DataFrame(index=dates, columns=['account_value'])

    for date in equity_df.index.date:  # Convert index to datetime.date for comparison
        daily_margin = sum(
            df[(df['entry_time'].dt.date <= date)]['margin'].sum() - 
            df[(df['exit_time'].dt.date <= date)]['margin'].sum() 
            for df in strategy_data.values()
        )
        equity_df.at[date, 'account_value'] = CAPITAL - daily_margin

    return equity_df


def main():
    # st.title("Strategy Performance Dashboard")

    # Read and process data
    file_path = r"C:\Users\user\Desktop\Kaas\StrategyAdmin\userexcel\omkarhegde_new.xlsx"
    strategy_data = read_strategy_logs(file_path)
    strategy_data = calculate_margins(strategy_data)

    # Risk Adjustment Section
    # with st.sidebar:
    #     st.header("Adjust Risk Parameters")
    #     # Assuming risk values are floats, make sure all slider components are floats too
    #     for strategy, default_risk in risk_profile.items():
    #         risk_profile[strategy] = st.slider(f"{strategy} Risk",
    #                                         min_value=0.0,  # float
    #                                         max_value=500000.0,  # float
    #                                         value=float(default_risk),  # ensuring default_risk is float
    #                                         step=0.01)  # Define the step for float values


    # Recalculate margins and equity chart based on updated risk
    # strategy_data = calculate_margins(strategy_data)
    equity_df = calculate_equity(strategy_data)
    print(equity_df)
    
    # # Before sending the DataFrame to Streamlit for rendering
    # equity_df.index = equity_df.index.strftime('%Y-%m-%d')  # Convert index to string if it's date
    # equity_df.reset_index(inplace=True)  # Resetting index to move date from index to a column

    # # Optionally convert all date columns to strings
    # equity_df['date'] = equity_df['date'].astype(str)

    # Display Equity Chart
    # st.line_chart(equity_df['account_value'])

if __name__ == "__main__":
    main()
