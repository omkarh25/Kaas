import pandas as pd
import os
import requests
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class PaymentProcessor:
    def __init__(self, excel_file, sheet_name):
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.df = pd.read_excel(excel_file, sheet_name=sheet_name)
        self.today = datetime.today().date()

    def filter_payments(self):
        self.df['Date'] = pd.to_datetime(self.df['Date']).dt.date
        self.df['Paid'] = self.df['Paid'].astype(str).str.lower() == 'Yes'
        
        self.today_payments = self.df[(self.df['Date'] == self.today) & (~self.df['Paid'])]
        self.past_due_payments = self.df[(self.df['Date'] < self.today) & (~self.df['Paid'])]
        self.upcoming_week_payments = self.df[(self.df['Date'] > self.today) & (self.df['Date'] <= self.today + timedelta(days=7)) & (~self.df['Paid'])]

    def generate_markdown(self, output_file):
        with open(output_file, 'w') as f:
            f.write("# Transactions\n\n")
            f.write("## Today's Payments\n")
            f.write(self._generate_todo_list(self.today_payments))
            f.write("\n## Past Due Payments\n")
            f.write(self._generate_todo_list(self.past_due_payments))
            f.write("\n## Upcoming Week Payments\n")
            f.write(self._generate_todo_list(self.upcoming_week_payments))

    def _generate_todo_list(self, df):
        todo_list = df.apply(lambda row: f"- [ ] {', '.join(row.astype(str))}", axis=1).tolist()
        return "\n".join(todo_list) + "\n"

    def update_excel_from_markdown(self, markdown_file):
        with open(markdown_file, 'r') as f:
            content = f.read()
        
        # Extract checked items
        checked_items = re.findall(r'- \[x\] (.+)', content)
        
        # Update the DataFrame
        for item in checked_items:
            values = item.split(', ')
            if len(values) > 1:
                index = int(values[0]) - 1  # Assuming the first value is the index
                self.df.at[index, 'Paid'] = 'Yes'
        
        # Load the existing workbook
        book = load_workbook(self.excel_file)
        
        # Remove the existing sheet
        if self.sheet_name in book.sheetnames:
            del book[self.sheet_name]
        
        # Create a new sheet
        sheet = book.create_sheet(self.sheet_name)
        
        # Write the updated DataFrame to the new sheet
        for r in dataframe_to_rows(self.df, index=False, header=True):
            sheet.append(r)
        
        # Save the workbook
        book.save(self.excel_file)

def excel_to_markdown(excel_file, sheet_name, output_file):
    processor = PaymentProcessor(excel_file, sheet_name)
    processor.filter_payments()
    processor.generate_markdown(output_file)

def update_excel_from_markdown(excel_file, sheet_name, markdown_file):
    processor = PaymentProcessor(excel_file, sheet_name)
    processor.update_excel_from_markdown(markdown_file)

# Usage
excel_file = 'Kaas.xlsx'
sheet_name = 'FreedomBlast(Future)'
output_file = '/Users/omkar/Documents/Obsidian Vault/4_DayToDay/02-Sep-24.md'
markdown_file = '/Users/omkar/Documents/Obsidian Vault/4_DayToDay/02-Sep-24.md'

# excel_to_markdown(excel_file, sheet_name, output_file)
update_excel_from_markdown(excel_file, sheet_name, markdown_file)
# TODO write checked items to a different sheet to avoid conflicts with NIkhil